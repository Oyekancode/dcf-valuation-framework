"""
Excel Export Module for DCF Valuations
======================================
Export DCF analysis to professionally formatted Excel workbooks
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict
from dcf_valuation import DCFValuation, CompanyData, ValuationAssumptions
import os


class ExcelExporter:
    """Export DCF valuation to Excel with professional formatting"""
    
    def __init__(self, dcf: DCFValuation, results: Dict):
        self.dcf = dcf
        self.results = results
        self.wb = Workbook()
        
        # Define styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_header_style(self, cell):
        """Apply header styling to a cell"""
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border
    
    def _apply_subheader_style(self, cell):
        """Apply subheader styling to a cell"""
        cell.fill = self.subheader_fill
        cell.font = self.subheader_font
        cell.border = self.border
    
    def _apply_highlight_style(self, cell):
        """Apply highlight styling to a cell"""
        cell.fill = self.highlight_fill
        cell.border = self.border
        cell.font = Font(bold=True)
    
    def create_summary_sheet(self):
        """Create executive summary sheet"""
        ws = self.wb.active
        ws.title = "Executive Summary"
        
        row = 1
        
        # Title
        ws[f'A{row}'] = f"DCF VALUATION: {self.dcf.company.company_name}"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Company Information
        ws[f'A{row}'] = "COMPANY INFORMATION"
        self._apply_subheader_style(ws[f'A{row}'])
        row += 1
        
        company_data = [
            ["Ticker", self.dcf.company.ticker],
            ["Company Name", self.dcf.company.company_name],
            ["Current Stock Price", f"${self.dcf.company.stock_price:.2f}"],
            ["Shares Outstanding (M)", f"{self.dcf.company.shares_outstanding:,.0f}"],
            ["Market Cap ($M)", f"${self.dcf.company.market_cap():,.0f}"],
            ["Current FCF ($M)", f"${self.dcf.company.current_fcf:,.0f}"],
            ["Net Debt ($M)", f"${self.dcf.company.net_debt:,.0f}"]
        ]
        
        for label, value in company_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Assumptions
        ws[f'A{row}'] = "VALUATION ASSUMPTIONS"
        self._apply_subheader_style(ws[f'A{row}'])
        row += 1
        
        assumptions_data = [
            ["Revenue Growth Rate", f"{self.dcf.assumptions.revenue_growth_rate*100:.1f}%"],
            ["Terminal Growth Rate", f"{self.dcf.assumptions.terminal_growth_rate*100:.1f}%"],
            ["WACC (Discount Rate)", f"{self.dcf.assumptions.wacc*100:.1f}%"],
            ["Projection Period", f"{self.dcf.assumptions.projection_years} years"]
        ]
        
        for label, value in assumptions_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Valuation Results
        ws[f'A{row}'] = "VALUATION RESULTS"
        self._apply_subheader_style(ws[f'A{row}'])
        row += 1
        
        results_data = [
            ["Enterprise Value ($M)", f"${self.results['enterprise_value']:,.0f}"],
            ["Less: Net Debt ($M)", f"${self.dcf.company.net_debt:,.0f}"],
            ["Equity Value ($M)", f"${self.results['equity_value']:,.0f}"],
            ["", ""],
            ["Fair Value Per Share", f"${self.results['fair_value_per_share']:.2f}"],
            ["Current Market Price", f"${self.results['current_price']:.2f}"],
            ["Upside/(Downside)", f"{self.results['upside_downside_pct']:+.1f}%"],
            ["Assessment", self.results['assessment']]
        ]
        
        for label, value in results_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if label in ["Fair Value Per Share", "Assessment"]:
                self._apply_highlight_style(ws[f'A{row}'])
                self._apply_highlight_style(ws[f'B{row}'])
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def create_fcf_projections_sheet(self):
        """Create FCF projections sheet"""
        ws = self.wb.create_sheet("FCF Projections")
        
        # Create DataFrame
        years = list(range(1, self.dcf.assumptions.projection_years + 1))
        df = pd.DataFrame({
            'Year': years,
            'FCF ($M)': self.results['fcf_projections'],
            'Growth Rate': [self.dcf.assumptions.revenue_growth_rate] * len(years),
            'Discount Factor': self.results['discount_factors'],
            'PV of FCF ($M)': self.results['pv_fcf']
        })
        
        # Write header
        ws['A1'] = "FREE CASH FLOW PROJECTIONS"
        ws['A1'].font = Font(size=12, bold=True)
        
        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 3:  # Header row
                    self._apply_header_style(cell)
                else:
                    cell.border = self.border
                    
                    # Format numbers
                    if c_idx in [2, 5]:  # FCF and PV columns
                        cell.number_format = '$#,##0'
                    elif c_idx in [3, 4]:  # Growth and discount factor
                        cell.number_format = '0.0%' if c_idx == 3 else '0.0000'
        
        # Add terminal value row
        last_row = 3 + len(df) + 2
        ws[f'A{last_row}'] = "Terminal Value"
        ws[f'B{last_row}'] = self.results['terminal_value']
        ws[f'B{last_row}'].number_format = '$#,##0'
        self._apply_highlight_style(ws[f'A{last_row}'])
        
        ws[f'A{last_row+1}'] = "PV of Terminal Value"
        ws[f'B{last_row+1}'] = self.results['pv_terminal_value']
        ws[f'B{last_row+1}'].number_format = '$#,##0'
        self._apply_highlight_style(ws[f'A{last_row+1}'])
        
        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 18
    
    def create_sensitivity_sheet(self):
        """Create sensitivity analysis sheet"""
        ws = self.wb.create_sheet("Sensitivity Analysis")
        
        ws['A1'] = "SENSITIVITY ANALYSIS: WACC vs Growth Rate"
        ws['A1'].font = Font(size=12, bold=True)
        
        # WACC sensitivity
        wacc_values = [self.dcf.assumptions.wacc - 0.02, 
                      self.dcf.assumptions.wacc - 0.01,
                      self.dcf.assumptions.wacc,
                      self.dcf.assumptions.wacc + 0.01,
                      self.dcf.assumptions.wacc + 0.02]
        
        growth_values = [self.dcf.assumptions.revenue_growth_rate - 0.05,
                        self.dcf.assumptions.revenue_growth_rate - 0.025,
                        self.dcf.assumptions.revenue_growth_rate,
                        self.dcf.assumptions.revenue_growth_rate + 0.025,
                        self.dcf.assumptions.revenue_growth_rate + 0.05]
        
        # Create sensitivity table
        row = 3
        col = 2
        
        # Header row (growth rates)
        ws.cell(row=row, column=1, value="WACC \\ Growth")
        self._apply_header_style(ws.cell(row=row, column=1))
        
        for growth in growth_values:
            ws.cell(row=row, column=col, value=f"{growth*100:.1f}%")
            self._apply_header_style(ws.cell(row=row, column=col))
            col += 1
        
        # Data rows
        row += 1
        for wacc in wacc_values:
            ws.cell(row=row, column=1, value=f"{wacc*100:.1f}%")
            self._apply_subheader_style(ws.cell(row=row, column=1))
            
            col = 2
            for growth in growth_values:
                # Calculate fair value with these parameters
                from dcf_valuation import ValuationAssumptions, DCFValuation
                
                temp_assumptions = ValuationAssumptions(
                    revenue_growth_rate=growth,
                    fcf_margin=self.dcf.assumptions.fcf_margin,
                    terminal_growth_rate=self.dcf.assumptions.terminal_growth_rate,
                    wacc=wacc,
                    projection_years=self.dcf.assumptions.projection_years
                )
                
                temp_dcf = DCFValuation(self.dcf.company, temp_assumptions)
                temp_results = temp_dcf.perform_valuation()
                
                cell = ws.cell(row=row, column=col, 
                             value=temp_results['fair_value_per_share'])
                cell.number_format = '$#,##0.00'
                cell.border = self.border
                
                # Highlight base case
                if abs(wacc - self.dcf.assumptions.wacc) < 0.001 and \
                   abs(growth - self.dcf.assumptions.revenue_growth_rate) < 0.001:
                    self._apply_highlight_style(cell)
                
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 7):
            ws.column_dimensions[chr(64 + col)].width = 12
    
    def export(self, filename: str):
        """
        Export to Excel file
        
        Args:
            filename: Output filename (should end in .xlsx)
        """
        self.create_summary_sheet()
        self.create_fcf_projections_sheet()
        self.create_sensitivity_sheet()
        
        self.wb.save(filename)
        print(f"✓ Excel report exported to: {filename}")
        return filename


def export_to_excel(dcf: DCFValuation, results: Dict, filename: str) -> str:
    """
    Convenience function to export DCF to Excel
    
    Args:
        dcf: DCFValuation object
        results: Valuation results dictionary
        filename: Output filename
    
    Returns:
        Path to exported file
    """
    exporter = ExcelExporter(dcf, results)
    return exporter.export(filename)
